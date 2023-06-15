from tkinter import *
from tkinter.ttk import Combobox
import  tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

#Code dealing with excel work

file=pathlib.Path('Backened_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="Contact"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"
    
    file.save('Backened_data.xlsx')


#Functions

def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=ageValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)


    file=openpyxl.load_workbook('Backened_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save(r'Backened_data.xlsx')
    messagebox.showinfo('info', 'details added!')

    nameValue.set('')
    contactValue.set('')
    ageValue.set('')
    addressEntry.delete(1.0,END)



def clear():
    nameValue.set('')
    contactValue.set('')
    ageValue.set('')
    addressEntry.delete(1.0,END)

#root fundamentals

root=Tk()
root.title("Data entery")
root.geometry('800x400+300+200')
root.resizable(True,True)
root.configure(bg="#326273")


#icon
icon_image=PhotoImage(file='mufasa.png')
root.iconphoto(False, icon_image)


#heading
Label(root,text="Please fill out this entery form: ",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

# #label
Label(root,text=("Name"), font=23, bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text=("Contact Number"), font=23, bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text=("Age"), font=23, bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text=("Gender"), font=23, bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text=("Address"), font=23, bg="#326273",fg="#fff").place(x=50,y=250)

#entry

nameValue=StringVar()
contactValue=StringVar()
ageValue=StringVar()

nameEntry= Entry(root, textvariable=nameValue, width=15, bd=2, font=20)
contactEntry=Entry(root, textvariable=contactValue, width=45, bd=2, font=20)
ageEntry=Entry(root, textvariable=ageValue, width=15, bd=2, font=20)

#Gender
gender_combobox=Combobox(root,values=['Male','Female'], font='arial 14', state='r', width=14)
gender_combobox.place(x=440,y=200)
gender_combobox.set('Male')

#address
addressEntry = Text(root,font=("arial"), width=50, height=4, bd=2)


nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250)

#button

Button(root,text="Submit",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Exit",bg="#326273",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)





root.mainloop()
